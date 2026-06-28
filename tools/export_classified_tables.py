from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Iterable

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


SUMMARY_FIELDS = [
    "category",
    "table_name",
    "match_status",
    "packet_asset_name",
    "packet_asset_path",
    "row_count",
    "column_count",
    "text_count",
    "rel_path",
    "json_output",
    "csv_output",
]

DEFAULT_EXCLUDE_CATEGORIES = {"gameplay"}


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export parsed named Packet tables into business-oriented JSON/CSV buckets."
    )
    parser.add_argument("--table-probe", required=True, help="Directory containing table_bins.csv and tables_json/.")
    parser.add_argument("--out", required=True, help="Output directory for classified tables.")
    parser.add_argument(
        "--report-root",
        default="",
        help="Optional root path to record in CSV/JSON/XLSX output paths when files are staged elsewhere.",
    )
    parser.add_argument(
        "--hide-physical-out",
        action="store_true",
        help="Do not record the staging output directory in summary.json when --report-root is used.",
    )
    parser.add_argument("--table-texts", help="Optional table_texts.csv or table_texts directory for text counts.")
    parser.add_argument("--exclude-gameplay", action="store_true", default=True, help="Skip tables already covered by gameplay export. Default: on.")
    parser.add_argument("--include-gameplay", action="store_true", help="Include gameplay tables too.")
    parser.add_argument("--table-regex", help="Only include table names matching this regex.")
    parser.add_argument("--category", action="append", default=[], help="Only include selected category. Repeatable.")
    parser.add_argument("--json-only", action="store_true", help="Only write JSON files.")
    parser.add_argument("--csv-only", action="store_true", help="Only write CSV files.")
    parser.add_argument("--xlsx", action="store_true", help="Also write classified_tables.xlsx when openpyxl is installed.")
    parser.add_argument("--max-rows", type=int, default=0, help="Maximum rows per exported table. 0 means all rows.")
    return parser.parse_args(argv)


def lower(value: str) -> str:
    return value.lower()


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._") or "unknown"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def normalize_rel_path(value: str) -> str:
    return value.replace("\\", "/").strip("/")


def table_json_path(tables_json_root: Path, rel_path: str) -> Path:
    return tables_json_root / f"{normalize_rel_path(rel_path)}.json"


def resolve_table_texts(path_text: str | None) -> Path | None:
    if not path_text:
        return None
    path = Path(path_text).expanduser().resolve()
    if path.is_dir():
        path = path / "table_texts.csv"
    return path if path.exists() else None


def load_text_counts(path: Path | None) -> Counter[str]:
    counts: Counter[str] = Counter()
    if path is None:
        return counts
    for row in read_csv(path):
        table_name = row.get("table_name", "")
        if table_name:
            counts[table_name] += 1
    return counts


def is_gameplay_table(name: str) -> bool:
    value = lower(name)
    if name.startswith("Table") and any(
        marker in value
        for marker in (
            "accessory",
            "ammo",
            "battle",
            "buff",
            "bullet",
            "character",
            "core",
            "damage",
            "debuff",
            "skill",
            "talent",
            "timeline",
        )
    ):
        return True
    if name in {
        "UiTableAccessory",
        "UiTableAccessorySet",
        "UiTableCharacterIndex",
        "UiTableCharacterParameter",
        "UiTableCommonFreeCharacter",
        "UiTableCore",
        "UiTableCoreType",
        "UiTableGroupSkill",
        "UiTableGroupSkillBranch",
        "UiTableRole",
        "UiTableRoleArchives",
        "UiTableRoleAttributeMap",
        "UiTableRoleProficiencyMain",
        "UiTableRoleRunesGlobal",
        "UiTableRoleRunesMain",
        "UiTableRoleSceneCamOffset",
        "UiTableRoleSkillIcon",
        "UiTableRoleType",
        "UiTableTechniqueTrainingRole",
    }:
        return True
    return False


def classify_table(name: str, fields: str, packet: str) -> str:
    name_lower = name.lower()
    value = f"{name} {fields} {packet}".lower()

    if is_gameplay_table(name):
        return "gameplay"
    if "equipment" in name_lower or "accessoryslotindex" in name_lower or "statusbar" in name_lower or "csbar" in name_lower:
        return "equipment_loadout"
    if "recharge" in name_lower or "mall" in name_lower or "shop" in name_lower or "monthcard" in name_lower or "lottery" in name_lower:
        return "shop_monetization"
    if name.startswith("UiTableActivity") or "mission" in name_lower or "achievement" in name_lower or "survey" in name_lower or name_lower in {"uitablematchtype", "uitabledailymission", "uitablerankweeklymission"}:
        return "activity_mission"
    if "item" in name_lower or "prop" in name_lower or "gift" in name_lower or "virtual" in name_lower or "headicon" in name_lower or "headlayout" in name_lower or "title" in name_lower or "namecard" in name_lower or "journeyofstarprogress" in name_lower:
        return "items_economy"
    if "jump" in name_lower or "tutorial" in name_lower or "guide" in name_lower or "loading" in name_lower or "gainway" in name_lower:
        return "navigation_tutorial"
    if "match" in name_lower or "rank" in name_lower or "map" in name_lower or "settlement" in name_lower or "pve" in name_lower or "journeyofstar" in name_lower:
        return "match_rank_battle_ui"
    if "role" in name_lower or "story" in name_lower or "painting" in name_lower or "settinghelp" in name_lower:
        return "profile_cosmetic"
    if "spine" in name_lower or "meme" in name_lower or "asset" in name_lower or "techniquetrainingchapters" in name_lower:
        return "visual_refs"
    if "behavior" in name_lower or "global" in name_lower or "package" in name_lower or "sundry" in name_lower:
        return "system_global"
    if "communication" in name_lower or "banpick" in name_lower or "replay" in name_lower:
        return "social_communication"
    if "area" in name_lower or "currency" in value or "language" in value or "recomendsetting" in name_lower or "recommendsetting" in name_lower:
        return "settings_region"
    if "activity" in value or "mission" in value or "achievement" in value:
        return "activity_mission"
    if "item" in value or "prop" in value or "gift" in value:
        return "items_economy"
    return "review"


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float, str)):
        return str(value)
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def collect_fields(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return fields


def write_table_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = collect_fields(rows)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: stringify_cell(row.get(field)) for field in fields})


def write_summary_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=SUMMARY_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def iter_selected_tables(args: argparse.Namespace) -> Iterable[tuple[dict[str, str], str, Path]]:
    table_probe = Path(args.table_probe).expanduser().resolve()
    table_report = table_probe / "table_bins.csv"
    tables_json_root = table_probe / "tables_json"
    if not table_report.exists():
        raise SystemExit(f"table_bins.csv not found: {table_report}")
    if not tables_json_root.exists():
        raise SystemExit(f"tables_json directory not found: {tables_json_root}")

    categories = set(args.category)
    table_pattern = re.compile(args.table_regex, re.IGNORECASE) if args.table_regex else None
    exclude_gameplay = bool(args.exclude_gameplay and not args.include_gameplay)

    for row in read_csv(table_report):
        table_name = row.get("table_name", "")
        if row.get("status") != "parsed" or not table_name:
            continue
        if table_pattern and not table_pattern.search(table_name):
            continue
        category = classify_table(table_name, row.get("field_names", ""), row.get("packet_asset_name", ""))
        if exclude_gameplay and category == "gameplay":
            continue
        if categories and category not in categories:
            continue
        source_path = table_json_path(tables_json_root, row.get("rel_path", ""))
        if not source_path.exists():
            continue
        yield row, category, source_path


def write_xlsx(path: Path, summary_rows: list[dict[str, str]], exported_tables: dict[str, list[dict[str, Any]]]) -> None:
    try:
        from openpyxl.workbook import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("openpyxl is required for --xlsx. Run: uv sync, then use uv run python ...") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Index"
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    ws.append(SUMMARY_FIELDS)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in summary_rows:
        ws.append([row.get(field, "") for field in SUMMARY_FIELDS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for index, field in enumerate(SUMMARY_FIELDS, start=1):
        width = min(max(len(field), 14), 48)
        ws.column_dimensions[get_column_letter(index)].width = width

    for table_name, rows in list(exported_tables.items())[:30]:
        sheet = wb.create_sheet(safe_name(table_name)[:31])
        dict_rows = [row for row in rows if isinstance(row, dict)]
        fields = collect_fields(dict_rows)
        sheet.append(fields)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in dict_rows[:5000]:
            sheet.append([stringify_cell(row.get(field)) for field in fields])
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
        for index, field in enumerate(fields, start=1):
            sample_width = max([len(field)] + [len(stringify_cell(item.get(field))) for item in dict_rows[:80]])
            sheet.column_dimensions[get_column_letter(index)].width = min(max(sample_width + 2, 10), 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def export_tables(args: argparse.Namespace) -> dict[str, Any]:
    if args.json_only and args.csv_only:
        raise SystemExit("--json-only and --csv-only cannot be used together.")

    out_root = Path(args.out).expanduser().resolve()
    report_root = Path(args.report_root).expanduser().resolve() if args.report_root else out_root
    text_counts = load_text_counts(resolve_table_texts(args.table_texts))
    summary_rows: list[dict[str, str]] = []
    exported_tables: dict[str, list[dict[str, Any]]] = {}
    category_counts: Counter[str] = Counter()
    row_total = 0

    for row, category, source_path in iter_selected_tables(args):
        table_name = row.get("table_name", "")
        table_rows = json.loads(source_path.read_text(encoding="utf-8"))
        if not isinstance(table_rows, list):
            continue
        if args.max_rows:
            table_rows = table_rows[: args.max_rows]

        file_stem = safe_name(table_name)
        json_output = out_root / category / "json" / f"{file_stem}.json"
        csv_output = out_root / category / "csv" / f"{file_stem}.csv"
        report_json_output = report_root / category / "json" / f"{file_stem}.json"
        report_csv_output = report_root / category / "csv" / f"{file_stem}.csv"

        if not args.csv_only:
            json_output.parent.mkdir(parents=True, exist_ok=True)
            json_output.write_text(json.dumps(table_rows, ensure_ascii=False, indent=2), encoding="utf-8")
        if not args.json_only:
            write_table_csv(csv_output, [item for item in table_rows if isinstance(item, dict)])

        exported_tables[table_name] = [item for item in table_rows if isinstance(item, dict)]
        category_counts[category] += 1
        row_total += len(table_rows)
        summary_rows.append(
            {
                "category": category,
                "table_name": table_name,
                "match_status": row.get("match_status", ""),
                "packet_asset_name": row.get("packet_asset_name", ""),
                "packet_asset_path": row.get("packet_asset_path", ""),
                "row_count": row.get("row_count", ""),
                "column_count": row.get("column_count", ""),
                "text_count": str(text_counts[table_name]),
                "rel_path": row.get("rel_path", ""),
                "json_output": str(report_json_output) if not args.csv_only else "",
                "csv_output": str(report_csv_output) if not args.json_only else "",
            }
        )

    summary_rows.sort(key=lambda item: (item["category"], item["table_name"]))
    write_summary_csv(out_root / "tables.csv", summary_rows)
    summary = {
        "table_probe": str(Path(args.table_probe).expanduser().resolve()),
        "out": str(report_root),
        "physical_out": "" if args.hide_physical_out else (str(out_root) if report_root != out_root else ""),
        "tables": len(summary_rows),
        "rows": row_total,
        "by_category": dict(sorted(category_counts.items())),
        "write_json": not args.csv_only,
        "write_csv": not args.json_only,
        "write_xlsx": bool(args.xlsx),
    }
    (out_root / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.xlsx:
        write_xlsx(out_root / "classified_tables.xlsx", summary_rows, exported_tables)

    (out_root / "README.md").write_text(
        "\n".join(
            [
                "# Classified Tables Export",
                "",
                "This directory contains parsed Packet tables grouped by business usage.",
                "",
                "- `tables.csv`: index of exported tables, categories, source Packet paths, and output files.",
                "- `<category>/json/*.json`: full structured rows.",
                "- `<category>/csv/*.csv`: Excel-readable UTF-8 with BOM tables.",
                "- `classified_tables.xlsx`: optional workbook when exported with `--xlsx`.",
                "",
                "Gameplay/battle tables are exported separately by `tools/export_gameplay_tables.py` unless `--include-gameplay` is used.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    return summary


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    summary = export_tables(args)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
