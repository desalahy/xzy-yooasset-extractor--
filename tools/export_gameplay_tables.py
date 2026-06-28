from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter
from pathlib import Path
from typing import Any, Iterable

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

TOOLS_ROOT = Path(__file__).resolve().parent
if str(TOOLS_ROOT) not in sys.path:
    sys.path.insert(0, str(TOOLS_ROOT))

from organize_exports import infer_table_domain


DEFAULT_DOMAINS = {"characters", "skills"}
DEFAULT_SUBDOMAINS = {
    "base",
    "stats",
    "core",
    "attributes",
    "accessory",
    "camera",
    "profile",
    "voice",
    "cooldown",
    "damage",
    "skill",
    "skill_tree",
    "skill_ui",
    "runes",
    "proficiency",
    "technique",
    "talent",
}

DIRECT_GAMEPLAY_PATTERNS = (
    "ammo",
    "battle",
    "buff",
    "bullet",
    "character",
    "core",
    "damage",
    "debuff",
    "skill",
    "talent",
    "timeline",
)

SUMMARY_FIELDS = [
    "business_domain",
    "business_subdomain",
    "confidence",
    "table_name",
    "match_status",
    "packet_asset_name",
    "packet_asset_path",
    "row_count",
    "column_count",
    "rel_path",
    "json_output",
    "csv_output",
]


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Export parsed gameplay/battle tables as standalone JSON and CSV files."
    )
    parser.add_argument(
        "--table-probe",
        required=True,
        help="Directory produced by probe_table_bins.py. Must contain table_bins.csv and tables_json/.",
    )
    parser.add_argument("--out", required=True, help="Output directory for gameplay table exports.")
    parser.add_argument(
        "--report-root",
        default="",
        help="Optional root path to record in CSV/JSON/XLSX output paths when files are staged elsewhere.",
    )
    parser.add_argument(
        "--hide-physical-out",
        action="store_true",
        help="Do not record the staging output directory in summary.json when --report-root is used.",
    )
    parser.add_argument(
        "--domain",
        action="append",
        default=[],
        help="Business domain to include, for example characters or skills. Repeatable. Defaults to gameplay domains.",
    )
    parser.add_argument(
        "--subdomain",
        action="append",
        default=[],
        help="Business subdomain to include, for example damage, cooldown, stats, core. Repeatable.",
    )
    parser.add_argument(
        "--table-regex",
        help="Additional regex filter against table_name. Useful for a narrow export such as Damage|Bullet|Ammo.",
    )
    parser.add_argument(
        "--include-ui-skill",
        action="store_true",
        help="Also include UiTableGroupSkill, UiTableRoleSkillIcon, and similar UI-facing skill tables.",
    )
    parser.add_argument("--json-only", action="store_true", help="Only write JSON files, not CSV mirrors.")
    parser.add_argument("--csv-only", action="store_true", help="Only write CSV files, not JSON mirrors.")
    parser.add_argument("--xlsx", action="store_true", help="Also write gameplay_tables.xlsx when openpyxl is installed.")
    parser.add_argument("--max-rows", type=int, default=0, help="Maximum rows per exported table. 0 means all rows.")
    return parser.parse_args(argv)


def safe_name(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_.-]+", "_", value.strip())
    return cleaned.strip("._") or "unknown"


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def normalize_rel_path(value: str) -> str:
    return value.replace("\\", "/").strip("/")


def table_json_path(tables_json_root: Path, rel_path: str) -> Path:
    return tables_json_root / f"{normalize_rel_path(rel_path)}.json"


def is_direct_gameplay_name(table_name: str) -> bool:
    lower = table_name.lower()
    return table_name.startswith("Table") and any(pattern in lower for pattern in DIRECT_GAMEPLAY_PATTERNS)


def include_row(
    row: dict[str, str],
    domains: set[str],
    subdomains: set[str],
    table_pattern: re.Pattern[str] | None,
    include_ui_skill: bool,
) -> tuple[bool, str, str, str]:
    domain, subdomain, confidence, _reason = infer_table_domain(row)
    table_name = row.get("table_name", "")

    if table_pattern and not table_pattern.search(table_name):
        return False, domain, subdomain, confidence
    if domain in domains and (not subdomains or subdomain in subdomains):
        return True, domain, subdomain, confidence
    if include_ui_skill and domain == "ui" and "skill" in table_name.lower():
        return True, domain, subdomain, confidence
    if is_direct_gameplay_name(table_name):
        return True, domain, subdomain, confidence
    return False, domain, subdomain, confidence


def stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, float, str)):
        return str(value)
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def collect_fields(rows: list[dict[str, Any]]) -> list[str]:
    fields: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                fields.append(key)
    return fields


def write_table_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = collect_fields(rows)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: stringify_cell(row.get(field)) for field in fields})


def write_summary_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=SUMMARY_FIELDS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, summary_rows: list[dict[str, str]], exported_tables: dict[str, list[dict[str, Any]]]) -> None:
    try:
        from openpyxl.workbook import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as exc:
        raise RuntimeError("openpyxl is required for --xlsx. Run: uv sync, then use uv run python ...") from exc

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    index_sheet = wb.active
    index_sheet.title = "Index"
    index_sheet.append(SUMMARY_FIELDS)
    for cell in index_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in summary_rows:
        index_sheet.append([row.get(field, "") for field in SUMMARY_FIELDS])
    index_sheet.freeze_panes = "A2"
    index_sheet.auto_filter.ref = index_sheet.dimensions
    for col_idx, field in enumerate(SUMMARY_FIELDS, start=1):
        index_sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(len(field) + 2, 14), 48)

    for table_name, rows in exported_tables.items():
        sheet = wb.create_sheet(safe_name(table_name)[:31])
        fields = collect_fields(rows)
        sheet.append(fields)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows[:5000]:
            sheet.append([stringify_cell(row.get(field)) for field in fields])
        sheet.freeze_panes = "A2"
        if sheet.max_row and sheet.max_column:
            sheet.auto_filter.ref = sheet.dimensions
        for col_idx, field in enumerate(fields, start=1):
            sample_width = max([len(field)] + [len(stringify_cell(item.get(field))) for item in rows[:80]])
            sheet.column_dimensions[get_column_letter(col_idx)].width = min(max(sample_width + 2, 10), 60)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def iter_selected_tables(args: argparse.Namespace) -> Iterable[tuple[dict[str, str], str, str, str, Path]]:
    table_probe = Path(args.table_probe).expanduser().resolve()
    table_report = table_probe / "table_bins.csv"
    tables_json_root = table_probe / "tables_json"
    if not table_report.exists():
        raise SystemExit(f"table_bins.csv not found: {table_report}")
    if not tables_json_root.exists():
        raise SystemExit(f"tables_json directory not found: {tables_json_root}")

    domains = set(args.domain or DEFAULT_DOMAINS)
    subdomains = set(args.subdomain or DEFAULT_SUBDOMAINS)
    table_pattern = re.compile(args.table_regex, re.IGNORECASE) if args.table_regex else None

    for row in read_csv(table_report):
        if row.get("status") != "parsed":
            continue
        if not row.get("table_name"):
            continue
        selected, domain, subdomain, confidence = include_row(
            row,
            domains,
            subdomains,
            table_pattern,
            args.include_ui_skill,
        )
        if not selected:
            continue
        source_path = table_json_path(tables_json_root, row.get("rel_path", ""))
        if not source_path.exists():
            continue
        yield row, domain, subdomain, confidence, source_path


def export_tables(args: argparse.Namespace) -> dict[str, Any]:
    if args.json_only and args.csv_only:
        raise SystemExit("--json-only and --csv-only cannot be used together.")

    out_root = Path(args.out).expanduser().resolve()
    report_root = Path(args.report_root).expanduser().resolve() if args.report_root else out_root
    summary_rows: list[dict[str, str]] = []
    exported_tables: dict[str, list[dict[str, Any]]] = {}
    table_counts: Counter[str] = Counter()
    row_total = 0

    for row, domain, subdomain, confidence, source_path in iter_selected_tables(args):
        table_name = row.get("table_name", "")
        table_rows = json.loads(source_path.read_text(encoding="utf-8"))
        if not isinstance(table_rows, list):
            continue
        if args.max_rows:
            table_rows = table_rows[: args.max_rows]

        file_stem = safe_name(table_name)
        json_output = out_root / domain / subdomain / "json" / f"{file_stem}.json"
        csv_output = out_root / domain / subdomain / "csv" / f"{file_stem}.csv"
        report_json_output = report_root / domain / subdomain / "json" / f"{file_stem}.json"
        report_csv_output = report_root / domain / subdomain / "csv" / f"{file_stem}.csv"

        if not args.csv_only:
            json_output.parent.mkdir(parents=True, exist_ok=True)
            json_output.write_text(json.dumps(table_rows, ensure_ascii=False, indent=2), encoding="utf-8")
        if not args.json_only:
            dict_rows = [item for item in table_rows if isinstance(item, dict)]
            write_table_csv(csv_output, dict_rows)
        else:
            dict_rows = [item for item in table_rows if isinstance(item, dict)]

        exported_tables[table_name] = dict_rows
        row_total += len(table_rows)
        table_counts[f"{domain}/{subdomain}"] += 1
        summary_rows.append(
            {
                "business_domain": domain,
                "business_subdomain": subdomain,
                "confidence": confidence,
                "table_name": table_name,
                "match_status": row.get("match_status", ""),
                "packet_asset_name": row.get("packet_asset_name", ""),
                "packet_asset_path": row.get("packet_asset_path", ""),
                "row_count": row.get("row_count", ""),
                "column_count": row.get("column_count", ""),
                "rel_path": row.get("rel_path", ""),
                "json_output": str(report_json_output) if not args.csv_only else "",
                "csv_output": str(report_csv_output) if not args.json_only else "",
            }
        )

    write_summary_csv(out_root / "tables.csv", summary_rows)
    summary = {
        "table_probe": str(Path(args.table_probe).expanduser().resolve()),
        "out": str(report_root),
        "physical_out": "" if args.hide_physical_out else (str(out_root) if report_root != out_root else ""),
        "tables": len(summary_rows),
        "rows": row_total,
        "by_group": dict(sorted(table_counts.items())),
        "write_json": not args.csv_only,
        "write_csv": not args.json_only,
        "write_xlsx": bool(args.xlsx),
    }
    (out_root / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    if args.xlsx:
        write_xlsx(out_root / "gameplay_tables.xlsx", summary_rows, exported_tables)

    (out_root / "README.md").write_text(
        "\n".join(
            [
                "# Gameplay Tables Export",
                "",
                "This directory contains standalone exports for parsed gameplay, character, skill, and battle tables.",
                "",
                "## First Open",
                "",
                "- `tables.csv`: index of exported tables and source Packet paths.",
                "- `characters/*/json`: character-facing table JSON.",
                "- `characters/*/csv`: Excel-readable character-facing CSV.",
                "- `skills/*/json`: skill, ammo, damage, bullet, buff, and talent JSON.",
                "- `skills/*/csv`: Excel-readable skill and battle CSV.",
                "- `gameplay_tables.xlsx`: optional workbook when exported with `--xlsx`.",
                "",
                "Arrays are kept as JSON strings in CSV cells so the CSV remains one row per original table row.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )
    return summary


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    summary = export_tables(args)
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
